# -*- coding: utf-8 -*-
"""
Created on Wed Jun  7 16:22:21 2023

@author: UM98XA
"""

import time
import itertools
from GoogleNews import GoogleNews
import pandas as pd
from datetime import datetime
from dateutil.relativedelta import relativedelta
import os
import openpyxl

################ Support Functions  ################ 
# Search Function
def search_google(query, language, region, item_list = []):
    # Open googlenews session
    gn = GoogleNews(lang=language, region=region)
    gn.enableException(True)
    
    # Perform Queries 
    gn.search(query)
    
    # Get top 10 results
    results = gn.results()
    if results:
        for item in results:
            item_list.append([query, item['title'], item['desc'], item['link'], item['media'], item['datetime'], item['date'], language])
    return item_list

# Convert exported 'date' to 'datetime' for nl sources
def define_date_nl(date):
    months = {'jan':1,'feb':2,'mrt':3,'apr':4,'mei':5,'jun':6,'jul':7,'aug':8,'sep':9,'':9,'okt':10,'nov':11,'dec':12, '01':1, '02':2, '03':3, '04':4, '05':5, '06':6, '07':7, '08':8, '09':9, '10':10, '11':11, '12':12}
    try:
        if ' geleden' in date.lower():
            q = int(date.split()[-3])
            if 'minuten' in date.lower() or 'mins' in date.lower():
                return datetime.now() + relativedelta(minutes=-q)
            elif 'uur' in date.lower():
                return datetime.now() + relativedelta(hours=-q)
            elif 'dag' in date.lower():
                return datetime.now() + relativedelta(days=-q)
            elif 'we' in date.lower():
                return datetime.now() + relativedelta(days=-7*q)
            elif 'maand' in date.lower():
                return datetime.now() + relativedelta(months=-q)
        elif 'gisteren' in date.lower():
            return datetime.now() + relativedelta(days=-1)
        else:
            date_list = date.replace('/',' ').split(' ')
            if len(date_list) == 2 and date_list[0].isnumeric:
                date_list.append(datetime.now().year)
            elif len(date_list) == 2 and not date_list[0].isnumeric:
                date_list = ['1'] + date_list
            elif len(date_list) == 3:
                if date_list[0] == '':
                    date_list[0] = '1'
            return datetime(day=int(date_list[0]), month=months[date_list[1]], year=int(date_list[2]))
    except:
        print(f"!!! Can't recognize date: {date}")
        return float('nan')

# Append dataframe to existing 
def append_df_to_excel(df, excel_path):
    # Load and concat
    df_excel = pd.read_excel(excel_path)
    result = pd.concat([df_excel, df], ignore_index=True)
    
    # deduplicate and export
    #result = result.groupby(['Title', 'Short Description', 'Link', 'Media Outlet', 'Publish Date', 'Search Language', 'Fraud Type'])['Search Query'].apply(lambda x: '; '.join(x)).reset_index()
    result.to_excel(excel_path, index=False)

def log_runtime(project_path, search_queries):
    # Log Metrics #
    file_name = "Fraud Queries.xlsx"
    file_path = os.path.join(project_path, file_name)
    
    # Save to workbook
    wb = openpyxl.load_workbook(file_path)
    ws = wb["Run Time Log"]
    
    # Get Run info of last run 
    if ws.max_row == 1:
        current_run_index = 1
        
    else:
        dict_last_row = {ws.cell(1, i).value: ws.cell(ws.max_row, i).value for i in range(1, ws.max_column + 1)}
        if dict_last_row['Month'] != datetime.now().strftime("%B"):    
            current_run_index = 1
        else:
            current_run_index = dict_last_row['Monthly Run Index'] + 1
    
    # Set Run info of current run
    current_run = {
        'Run Time':datetime.now().strftime("%m/%d/%Y, %H:%M:%S"),
        'Year':datetime.now().strftime("%Y"),
        'Month':datetime.now().strftime("%B"),
        'Day':datetime.now().strftime("%A"),
        'Fraud Queries':", ".join(search_queries),
        'Monthly Run Index':current_run_index
    }
    
    # Get run info of current run
    ws.append(list(current_run.values()))
    wb.save(file_path)
    wb.close()
    return current_run_index

################ Main Function      ################
def export_googlenews(): 
    # Set parameters
    project_path = os.path.join(r"C:\Users", os.environ['UserName'], r"ING\NFR Data & Analytics - General\Fraud Dashboard\Google News Scraper")
    region = "NL"
    languages = ["en", "nl"]
    
    # Set queries:
    # Run max 5 query's per day (en + nl: 10 total), script is run every working day in the first week of the month
    # If you run to many queries on a given day Google may decide to block you for a day causing the script to fail
    df_search_queries = pd.read_excel(os.path.join(project_path, 'Fraud Queries.xlsx'), sheet_name="Schedule")
    search_queries = [[query, day] for query, day in zip(df_search_queries['Search Query'].tolist(), df_search_queries['Day'].tolist())]
    
    # Filter Todays Query's
    search_queries = [item[0] for item in search_queries if item[1] == datetime.now().strftime("%A")]
    
    # Log Run time
    current_run_index = log_runtime(project_path, search_queries)
        
    # Exit if no queries
    if not search_queries:
        print("Exit function: No queries for today")
        return current_run_index
    
    # Loop over each combination of options and queries: Search for each
    item_list = []
    i = 1
    for language, query in itertools.product(languages, search_queries):
        print(f"{i}: {language} - {query}")
        try:
            item_list = search_google(query, language, region, item_list)
        except Exception as e:
            if str(e) == "'NoneType' object is not iterable":
                pass
            else:
                print(e)
                raise Exception("HTTP Error!!!!")
            
        time.sleep(15) # Add delay between requests to prevent IP Blocking
        i += 1
        
    # Convert to dataframe
    df_data = pd.DataFrame(item_list, columns=['Search Query', 'Title', 'Short Description', 'Link', 'Media Outlet', 'Publish Datetime', 'Publish Date', 'Search Language'])
    
    # Convert dutch date formats + texts to datetime    
    df_data['Publish Datetime'] = df_data.apply(lambda row: define_date_nl(row['Publish Date']) if row['Search Language'] == "nl" else row['Publish Datetime'], axis=1)
    
    # Remove datetime elements into dates
    df_data['Publish Date'] = pd.to_datetime(df_data['Publish Datetime']).dt.date
    df_data.drop('Publish Datetime', axis=1, inplace=True)
    
    # Add Fraud type
    df_data = pd.merge(df_data, df_search_queries[['Search Query', 'Fraud Type']], on=['Search Query'])
    
    # Group duplicates
    #df_data = df_data.groupby(['Title', 'Short Description', 'Link', 'Media Outlet', 'Publish Date', 'Search Language', 'Fraud Type'])['Search Query'].apply(lambda x: '; '.join(x)).reset_index()
    
    # Remove old stories: within last 2 months
    df_data = df_data.loc[df_data['Publish Date'] >= datetime.date(datetime.now() + relativedelta(months=-2))]
    
    # Export to Excel or append to current export
    filepath = os.path.join(project_path, "Exports", datetime.now().strftime("%Y %b ") + "Fraud GoogleNews.xlsx")
    if not os.path.exists(filepath):
        # Create New Export
        print("create new file")
        df_data.to_excel(filepath, index=False)
    else:
        # Append to Existing dataset
        print("append to old file")
        append_df_to_excel(df_data, filepath)
    return current_run_index